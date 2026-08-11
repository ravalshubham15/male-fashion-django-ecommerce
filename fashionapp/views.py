from django.shortcuts import render, redirect, HttpResponse, get_object_or_404
from django.contrib.auth.models import User
from urllib3 import request 
from .models import userregister,brand,category,product,AddToCart,Order,OrderItem,Wishlist,Review,Payment
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.db.models import Sum, Avg
from django.contrib.auth.decorators import login_required
import random
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from django.utils import timezone
import json
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import os
from django.conf import settings
from reportlab.platypus import Image, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models.functions import TruncMonth
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from decimal import Decimal

# Create your views here.

def index(request):

    search = request.GET.get('search')

    if search:
        prod = product.objects.filter(name__icontains=search)
    else:
        prod = product.objects.all()

    brands = brand.objects.all()[:6]

    context = {
        'prod': prod,
        'brands': brands,
    }

    return render(request, 'index.html', context)

def adminsite(request):
    return render(request,'adminsite/index.html')

def registeruser(request):
    if request.method == 'POST':
        name= request.POST.get('txtname')
        username= request.POST.get('txtusername')
        email= request.POST.get('txtemail')
        password= request.POST.get('txtpassword')
        confirm_password= request.POST.get('txtconfirmpassword')
        gender= request.POST.get('txtgender')
        contact= request.POST.get('txtcontact')
        city= request.POST.get('txtcity')
        
        if password != confirm_password:
            messages.error(request, "PASSWORD DO NOT MATCH...")
            return render(request, 'registeruser.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, "USERNAME ALREADY EXISTS...")
            return render(request, 'registeruser.html')
        
        user= User.objects.create_user(username=username, email=email, password=password)
        userregister.objects.create(
            user=user,
            name=name,
            gender=gender,
            contact=contact,
            city=city,
        )
        messages.success(request, "REGISTRATION SUCCESSFUL. PLEASE LOGIN...")
        return redirect('userlogin')
    return render(request, 'registeruser.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('index')
        else:
            messages.error(request, "INVALID CREDENTIALS...")
    return render(request, 'userlogin.html')

def logout_view(request):
    logout(request)
    return redirect('index')

def home_view(request):
    return render(request, 'home.html')
             
def login_view(request):
    if request.method == 'POST':
        username = request.POST['txtusername']
        password = request.POST['txtpassword']
        user = authenticate(username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('index')
        else:
            messages.error(request, "INVALID CREDENTIALS...")
    return render(request, 'userlogin.html') 

def manageuser(request):
    data = userregister.objects.select_related("user").all().order_by("-id")
    return render(request,"adminsite/manageuser.html",{"data": data})

def adduser(request):
    if request.method == 'POST':
        name= request.POST.get('txtname')
        username= request.POST.get('txtusername')
        email= request.POST.get('txtemail')
        password= request.POST.get('txtpassword')
        confirm_password= request.POST.get('txtconfirmpassword')
        gender= request.POST.get('txtgender')
        contact= request.POST.get('txtcontact')
        city= request.POST.get('txtcity')
        
        if password != confirm_password:
            messages.error(request, "PASSWORD DO NOT MATCH...")
            return render(request, 'adminsite/adduser.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, "USERNAME ALREADY EXISTS...")
            return render(request, 'adminsite/adduser.html')
        
        user= User.objects.create_user(username=username, email=email, password=password)
        userregister.objects.create(
            user=user,
            name=name,
            gender=gender,
            contact=contact,
            city=city,
        )
        messages.success(request, "REGISTRATION SUCCESSFUL. PLEASE LOGIN...")
        return redirect('adminsite/manageuser.html')
    return render(request,'adminsite/adduser.html')

def deleteuser(request,uid):
    data=get_object_or_404(userregister,id=uid)
    data.user.delete()
    return redirect('manageuser')

def edituser(request, uid):

    data = get_object_or_404(userregister, id=uid)

    if request.method == "POST":

        data.name = request.POST.get("txtname")
        data.contact = request.POST.get("txtcontact")
        data.city = request.POST.get("txtcity")
        data.gender = request.POST.get("txtgender")

        if request.FILES.get("txtimage"):
            data.profile_image = request.FILES.get("txtimage")

        data.save()

        return redirect("manageuser")

    return render(request,
                  "adminsite/edituser.html",
                  {"data": data})

def addbrand(request):
    if request.method == "POST":
        bname = request.POST.get("txtbname")
        description = request.POST.get("txtdescription")
        image = request.FILES.get("txtimage")

        if bname and image:
            brand.objects.create(
                name=bname,
                description=description,
                image=image
            )
            return redirect("managebrand")

    return render(request, "adminsite/addbrand.html")

def managebrand(request):
    data=brand.objects.all()
    return render(request,'adminsite/managebrand.html',{'data':data})

def deletebrand(request,bid):
    data=get_object_or_404(brand,id=bid)
    data.delete()
    return redirect('managebrand')

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render

def editbrand(request, id):
    data = get_object_or_404(brand, id=id)

    if request.method == "POST":
        print("Current Brand ID:", data.id)
        print("Old Name:", data.name)

        name = request.POST.get("txtbname")
        print("New Name:", name)

        data.name = name
        data.description = request.POST.get("txtdescription")

        if request.FILES.get("txtimage"):
            data.image = request.FILES.get("txtimage")

            print("POST Name:", request.POST.get("txtbname"))
            print("Current DB Name:", data.name)

        data.save()

        return redirect("managebrand")

    return render(request, "adminsite/editbrand.html", {"data": data})

def addcategory(request):
    brands = brand.objects.all()

    if request.method == "POST":
        name = request.POST.get("txtcname")
        description = request.POST.get("txtdescription")
        brand_id = request.POST.get("txtbrand")

        if name and brand_id:
            category.objects.create(
                name=name,
                description=description,
                brand_id=brand_id
            )
            return redirect("managecategory")

    return render(request, "adminsite/addcategory.html", {
        "data": brands
    })

def managecategory(request):
    data=category.objects.select_related('brand').all()
    return render(request,'adminsite/managecategory.html', {'data':data})

def deletecategory(request,cid):
    data=get_object_or_404(category,id=cid)
    data.delete()
    return redirect('managecategory')

def editcategory(request, cid):
    data = get_object_or_404(category, id=cid)
    brands = brand.objects.all()

    if request.method == "POST":
        data.name = request.POST.get("txtcname")
        data.description = request.POST.get("txtdescription")
        data.brand_id = request.POST.get("txtbrand")

        data.save()

        return redirect("managecategory")

    return render(request, "adminsite/editcategory.html", {
        "data": data,
        "brands": brands
    })

def addproduct(request):
    brands = brand.objects.all()

    selected_brand_id = request.POST.get("brand")
    categories = category.objects.filter(brand_id=selected_brand_id) if selected_brand_id else []

    if request.method == "POST":

        name = request.POST.get("txtpname")
        selected_category_id = request.POST.get("category")
        price = request.POST.get("txtprice")
        description = request.POST.get("txtdescription")
        stock = request.POST.get("txtstock")
        image = request.FILES.get("txtimage")

        if not all([name, selected_brand_id, selected_category_id, price, stock, description, image]):
            messages.error(request, "Please fill all fields.")
        else:

            if product.objects.filter(name=name).exists():
                messages.warning(request, "Product already exists.")
            else:

                product.objects.create(
                    name=name,
                    brand_id=selected_brand_id,
                    category_id=selected_category_id,
                    price=price,
                    description=description,
                    stock=stock,
                    image=image,
                )

                messages.success(request, "Product added successfully.")
                return redirect("manageproduct")

    return render(request, "adminsite/addproduct.html", {
        "brands": brands,
        "categories": categories,
        "selected_brand_id": selected_brand_id,
    })
    
def manageproduct(request):
    data=product.objects.all()
    return render(request,'adminsite/manageproduct.html',{'data':data})

def editproduct(request, id):
    p = get_object_or_404(product, id=id)

    brands = brand.objects.all()

    if request.method == "POST":
        p.name = request.POST.get("txtpname")
        p.brand_id = request.POST.get("brand")
        p.category_id = request.POST.get("category")
        p.price = request.POST.get("txtprice")
        p.description = request.POST.get("txtdescription")
        p.stock = request.POST.get("txtstock")

        if request.FILES.get("txtimage"):
            p.image = request.FILES.get("txtimage")

        p.save()

        return redirect("manageproduct")

    categories = category.objects.filter(brand=p.brand)

    return render(request, "adminsite/editproduct.html", {
        "product": p,
        "brands": brands,
        "categories": categories,
    })

def deleteproduct(request,pid):
    data=get_object_or_404(product,id=pid)
    data.delete()
    return redirect('manageproduct')

def addtocart(request, pid):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    p = get_object_or_404(product, id=pid)

    cart_item, created = AddToCart.objects.get_or_create(
        user=request.user,
        product=p
    )

    if not created:
        cart_item.quantity += 1
        cart_item.save()

    return redirect('cart')

def cart(request):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    cart_items = AddToCart.objects.filter(user=request.user)

    total = 0

    for item in cart_items:
        item.subtotal = item.product.price * item.quantity
        total += item.subtotal

    return render(request, "shopping-cart.html", {
        "cart_items": cart_items,
        "total": total
    })

def removecart(request, cid):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    cart_item = get_object_or_404(AddToCart, id=cid, user=request.user)
    cart_item.delete()

    return redirect('cart')

def increase_quantity(request, cid):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    cart_item = get_object_or_404(AddToCart, id=cid, user=request.user)

    cart_item.quantity += 1
    cart_item.save()

    return redirect('cart')

def decrease_quantity(request, cid):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    cart_item = get_object_or_404(AddToCart, id=cid, user=request.user)

    if cart_item.quantity > 1:
        cart_item.quantity -= 1
        cart_item.save()
    else:
        cart_item.delete()

    return redirect('cart')

def place_order(request):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    cart_items = AddToCart.objects.filter(user=request.user)

    if not cart_items.exists():
        return redirect('cart')

    total = 0

    for item in cart_items:
        total += item.product.price * item.quantity

    order = Order.objects.create(
        user=request.user,
        total_amount=total
    )

    for item in cart_items:
        OrderItem.objects.create(
            order=order,
            product=item.product,
            quantity=item.quantity,
            price=item.product.price
        )

    cart_items.delete()

    return render(request, "order-success.html")

def my_orders(request):

    if not request.user.is_authenticated:
        return redirect("userlogin")

    orders = Order.objects.filter(
        user=request.user
    ).order_by("-order_date")

    order_data = []

    for order in orders:

        payment = Payment.objects.filter(order=order).first()

        order_data.append({
            "order": order,
            "payment": payment
        })

    return render(request, "my-orders.html", {
        "order_data": order_data
    })

def my_profile(request):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    profile, created = userregister.objects.get_or_create(
        user=request.user,
        defaults={
            "name": request.user.username,
            "contact": "",
            "city": "",
            "gender": ""
        }
    )

    total_orders = Order.objects.filter(user=request.user).count()
    total_cart = AddToCart.objects.filter(user=request.user).count()
    total_wishlist = Wishlist.objects.filter(user=request.user).count()

    context = {
        "profile": profile,
        "total_orders": total_orders,
        "total_cart": total_cart,
        "total_wishlist": total_wishlist,
    }

    return render(request, "my-profile.html", context)

def edit_profile(request):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    profile, created = userregister.objects.get_or_create(
        user=request.user,
        defaults={
            "name": request.user.username,
            "contact": "",
            "city": "",
            "gender": ""
        }
    )

    if request.method == "POST":
        profile.name = request.POST.get("name")
        profile.contact = request.POST.get("contact")
        profile.city = request.POST.get("city")
        profile.gender = request.POST.get("gender")

        if request.FILES.get("profile_image"):
            profile.profile_image = request.FILES["profile_image"]

        profile.save()

        return redirect("my_profile")

    return render(request, "edit-profile.html", {
        "profile": profile
    })

def change_password(request):
    if not request.user.is_authenticated:
        return redirect("userlogin")

    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)

        if form.is_valid():
            user = form.save()

            update_session_auth_hash(request, user)

            messages.success(request, "Password changed successfully.")

            return redirect("my_profile")

        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = PasswordChangeForm(request.user)

    return render(request, "change-password.html", {
        "form": form
    })

from .models import Review

def product_details(request, pid):

    product_data = get_object_or_404(product, id=pid)

    reviews = Review.objects.filter(
        product=product_data
    ).select_related("user").order_by("-created_at")

    average_rating = reviews.aggregate(
        Avg("rating")
    )["rating__avg"] or 0

    review_count = reviews.count()

    related_products = product.objects.filter(
        category=product_data.category
    ).exclude(id=product_data.id)[:4]

    return render(request, "product-details.html", {
        "product": product_data,
        "reviews": reviews,
        "average_rating": round(average_rating, 1),
        "review_count": review_count,
        "related_products": related_products,
    })

@login_required
def add_to_wishlist(request, pid):
    pro = get_object_or_404(product, id=pid)

    wishlist_item, created = Wishlist.objects.get_or_create(
        user=request.user,
        product=pro
    )

    if created:
        messages.success(request, "❤️ Product added to Wishlist!")
    else:
        messages.info(request, "⚠️ Product is already in your Wishlist!")

    return redirect('index')

@login_required
def wishlist(request):
    wish = Wishlist.objects.filter(user=request.user)

    cart_count = AddToCart.objects.filter(user=request.user).count()

    context = {
        'wish': wish,
        'cart_count': cart_count,
    }

    return render(request, 'wishlist.html', context)

def remove_wishlist(request, wid):
    item = get_object_or_404(Wishlist, id=wid, user=request.user)
    item.delete()
    return redirect('wishlist')

def add_review(request, pid):
    if request.method == "POST":

        pro = get_object_or_404(product, id=pid)

        rating = request.POST.get("rating")
        review = request.POST.get("review")

        Review.objects.create(
            product=pro,
            user=request.user,
            rating=rating,
            review=review
        )

        messages.success(request, "⭐ Review added successfully!")

    return redirect('product_details', pid=pid)

def delete_review(request, rid):

    review = get_object_or_404(Review, id=rid)

    # Only the review owner can delete it
    if review.user != request.user:
        messages.error(request, "You cannot delete this review.")
        return redirect("product_details", pid=review.product.id)

    pid = review.product.id
    review.delete()

    messages.success(request, "🗑️ Review deleted successfully!")

    return redirect("product_details", pid=pid)

def edit_review(request, rid):

    review = get_object_or_404(Review, id=rid)

    # Only the review owner can edit
    if review.user != request.user:
        messages.error(request, "You cannot edit this review.")
        return redirect("product_details", pid=review.product.id)

    if request.method == "POST":
        review.rating = request.POST.get("rating")
        review.review = request.POST.get("review")
        review.save()

        messages.success(request, "✅ Review updated successfully!")
        return redirect("product_details", pid=review.product.id)

    return render(request, "edit_review.html", {
        "review": review
    })

from django.utils import timezone
import random

def payment(request):

    if not request.user.is_authenticated:
        return redirect("userlogin")

    cart_items = AddToCart.objects.filter(user=request.user)

    if not cart_items.exists():
        messages.error(request, "Your cart is empty.")
        return redirect("cart")

    total = 0

    for item in cart_items:
        total += item.product.price * item.quantity

    if request.method == "POST":

        payment_method = request.POST.get("payment_method")

        payment_id = "PAY" + str(random.randint(100000, 999999))

        # Create Order
        order = Order.objects.create(
            user=request.user,
            total_amount=total,
            status="Pending"
        )

        # Create Order Items
        for item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )

        # Save Payment
        Payment.objects.create(
            user=request.user,
            order=order,
            payment_id=payment_id,
            amount=total,
            payment_method=payment_method,
            status="Success"
        )

        # Clear Cart
        cart_items.delete()

        return render(request, "payment_success.html", {
            "payment_id": payment_id,
            "amount": total,
            "payment_method": payment_method,
            "order": order,
        })

    return render(request, "payment.html", {
        "cart_items": cart_items,
        "total": total,
    })

@login_required
def download_invoice(request, order_id):

    order = Order.objects.get(id=order_id, user=request.user)
    order_items = OrderItem.objects.filter(order=order)
    payment = Payment.objects.filter(order=order).first()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{order.id}.pdf"'

    p = canvas.Canvas(response)
    p.setTitle(f"Invoice_{order.id}")

    logo_path = os.path.join(settings.BASE_DIR, "static", "img", "sr_logo.png")

    # ===========================
    # LOGO
    # ===========================

    if os.path.exists(logo_path):
        p.drawImage(
            logo_path,
            45,
            735,
            width=75,
            height=75,
            preserveAspectRatio=True,
            mask='auto'
        )

    # ===========================
    # COMPANY HEADER
    # ===========================

    p.setFont("Helvetica-Bold", 24)
    p.drawString(135, 810, "SR CLOTHING")

    p.setFont("Helvetica", 11)
    p.setFillColorRGB(.4,.4,.4)
    p.drawString(135,790,"Premium Men's Fashion Store")

    p.setFont("Helvetica",10)
    p.drawString(135,775,"Email : support@srclothing.com")
    p.drawString(135,760,"Phone : +91 98765 43210")

    p.setFillColorRGB(0,0,0)
    p.setFont("Helvetica-Bold",20)
    p.drawRightString(560,810,"INVOICE")

    p.line(40,745,560,745)

    # ===========================
    # CUSTOMER BOX
    # ===========================

    p.setFillColorRGB(.95,.95,.95)
    p.roundRect(40,600,520,120,8,fill=1)

    p.setFillColorRGB(0,0,0)

    y=690

    p.setFont("Helvetica-Bold",12)
    p.drawString(60,y,"Invoice No :")
    p.setFont("Helvetica",12)
    p.drawString(170,y,f"INV-{order.id}")

    y-=22

    p.setFont("Helvetica-Bold",12)
    p.drawString(60,y,"Order ID :")
    p.setFont("Helvetica",12)
    p.drawString(170,y,str(order.id))

    y-=22

    p.setFont("Helvetica-Bold",12)
    p.drawString(60,y,"Customer :")
    p.setFont("Helvetica",12)
    p.drawString(170,y,request.user.username)

    y-=22

    p.setFont("Helvetica-Bold",12)
    p.drawString(60,y,"Date :")
    p.setFont("Helvetica",12)
    p.drawString(
        170,
        y,
        timezone.localtime(order.order_date).strftime("%d-%m-%Y %I:%M %p")
    )

    # ===========================
    # PRODUCT TABLE
    # ===========================

    y=560

    p.line(40,y,560,y)

    y-=25

    p.setFont("Helvetica-Bold",12)

    p.drawString(60,y,"Product")
    p.drawString(330,y,"Qty")
    p.drawString(470,y,"Price")

    y-=10

    p.line(40,y,560,y)

    y-=25

    p.setFont("Helvetica",11)

    for item in order_items:

        p.drawString(60,y,item.product.name)

        p.drawString(340,y,str(item.quantity))

        p.drawRightString(540,y,f"Rs. {item.price}")

        y-=22

    p.line(40,y,560,y)

    y-=35

    # ===========================
    # TOTAL
    # ===========================

    p.setFont("Helvetica-Bold",16)

    p.drawRightString(
        540,
        y,
        f"Grand Total : Rs. {order.total_amount}"
    )

    y-=45

    # ===========================
    # PAYMENT
    # ===========================

    if payment:

        p.setFont("Helvetica-Bold",12)
        p.drawString(60,y,"Payment Details")

        y-=22

        p.setFont("Helvetica",11)

        p.drawString(60,y,f"Method : {payment.payment_method}")

        y-=18

        p.drawString(60,y,f"Status : {payment.status}")

        y-=18

        p.drawString(60,y,f"Payment ID : {payment.payment_id}")

        y-=40

    # ===========================
    # FOOTER
    # ===========================

    p.line(40,y,560,y)

    y-=30

    p.setFont("Helvetica-Bold",14)

    p.drawCentredString(
        300,
        y,
        "Thank You For Shopping With SR CLOTHING"
    )

    y-=18

    p.setFont("Helvetica",10)

    p.drawCentredString(
        300,
        y,
        "Generated by SR CLOTHING Management System"
    )

    p.save()

    return response

def order_details(request, order_id):  

    if not request.user.is_authenticated: return redirect("userlogin") 

    order = get_object_or_404( Order, id=order_id, user=request.user ) 

    order_items = OrderItem.objects.filter( order=order ).select_related("product") 

    subtotal = Decimal("0.00") 

    for item in order_items: 
        item.item_total = item.price * item.quantity 
        subtotal += item.item_total 
        payment = Payment.objects.filter( 
            order=order, 
            user=request.user ).first() 

        context = { 
            "order": order, 
            "order_items": order_items, 
            "payment": payment, 
            "subtotal": subtotal, } 

        return render( request, "order_details.html", context )

def cancel_order(request, order_id):

    order = Order.objects.get(id=order_id, user=request.user)

    if order.status == "Pending":
        order.status = "Cancelled"
        order.save()
        messages.success(request, "Your order has been cancelled successfully.")

    else:
        messages.error(request, "This order cannot be cancelled.")

    return redirect("my_orders")

@login_required
def admin_dashboard(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    total_users = User.objects.count()
    total_products = product.objects.count()
    total_orders = Order.objects.count()
    total_reviews = Review.objects.count()
    total_payments = Payment.objects.count()

    total_revenue = sum(
        order.total_amount
        for order in Order.objects.filter(status="Delivered")
    )

    recent_orders = (Order.objects.select_related("user").order_by("-order_date")[:5])
    recent_reviews = Review.objects.order_by("-created_at")[:5]

    context = {
        "total_users": total_users,
        "total_products": total_products,
        "total_orders": total_orders,
        "total_reviews": total_reviews,
        "total_payments": total_payments,
        "total_revenue": total_revenue,
        "recent_orders": recent_orders,
        "recent_reviews": recent_reviews,
        }

    return render(request, "adminsite/dashboard.html", context)

@login_required
def managepayment(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    payments = Payment.objects.select_related(
        "user",
        "order"
    ).order_by("-created_at")

    context = {
        "payments": payments
    }

    return render(
        request,
        "adminsite/payment.html",
        context
    )

@login_required
def managereview(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    reviews = Review.objects.select_related(
        "user",
        "product"
    ).order_by("-created_at")

    context = {
        "reviews": reviews
    }

    return render(
        request,
        "adminsite/review.html",
        context
    )

@login_required
def deletereview(request, rid):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    review = get_object_or_404(Review, id=rid)
    review.delete()

    return redirect("managereview")

@login_required
def manageorder(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    orders = Order.objects.select_related("user").order_by("-order_date")

    context = {
        "orders": orders
    }

    return render(
        request,
        "adminsite/manageorder.html",
        context
    )

@login_required
def updateorder(request, oid):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    order = get_object_or_404(Order, id=oid)

    if request.method == "POST":

        order.status = request.POST.get("status")
        order.save()

        return redirect("manageorder")

    context = {
        "order": order,
        "status_choices": Order.ORDER_STATUS
    }

    return render(
        request,
        "adminsite/updateorder.html",
        context
    )

@login_required
def deleteorder(request, oid):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    order = get_object_or_404(Order, id=oid)

    order.delete()

    return redirect("manageorder")

@login_required
def reports(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    total_users = User.objects.count()
    total_products = product.objects.count()
    total_orders = Order.objects.count()
    total_payments = Payment.objects.count()
    total_reviews = Review.objects.count()

    total_revenue = sum(
        order.total_amount
        for order in Order.objects.filter(status="Delivered")
    )

    context = {
        "total_users": total_users,
        "total_products": total_products,
        "total_orders": total_orders,
        "total_payments": total_payments,
        "total_reviews": total_reviews,
        "total_revenue": total_revenue,
    }

    return render(request, "adminsite/reports.html", context)

@login_required
def reports(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    # Dashboard Statistics
    total_users = User.objects.count()
    total_products = product.objects.count()
    total_orders = Order.objects.count()
    total_payments = Payment.objects.count()
    total_reviews = Review.objects.count()

    # Revenue
    total_revenue = (
        Order.objects.filter(status="Delivered")
        .aggregate(total=Sum("total_amount"))["total"] or 0
    )

    # Order Statistics
    today = timezone.now().date()

    today_orders = Order.objects.filter(order_date__date=today).count()
    pending_orders = Order.objects.filter(status="Pending").count()
    delivered_orders = Order.objects.filter(status="Delivered").count()
    cancelled_orders = Order.objects.filter(status="Cancelled").count()

    # Monthly Sales Chart
    monthly_sales = (
        Order.objects.filter(status="Delivered")
        .annotate(month=TruncMonth("order_date"))
        .values("month")
        .annotate(total=Sum("total_amount"))
        .order_by("month")
    )

    months = []
    sales = []

    for item in monthly_sales:
        months.append(item["month"].strftime("%b"))
        sales.append(float(item["total"]))

    # Pie Chart Data
    status_labels = json.dumps(["Pending", "Delivered", "Cancelled"])
    status_data = json.dumps([
        pending_orders,
        delivered_orders,
        cancelled_orders
    ])

    top_products = (
    OrderItem.objects
    .values("product__name")
    .annotate(
        total_sold=Sum("quantity"),
        revenue=Sum("price")
    )
    .order_by("-total_sold")[:5]
)

    latest_orders = (
    Order.objects
    .select_related("user")
    .order_by("-order_date")[:5]
)

    low_stock_products = product.objects.filter(stock__lte=5).order_by("stock")

    context = {
        "total_users": total_users,
        "total_products": total_products,
        "total_orders": total_orders,
        "total_payments": total_payments,
        "total_reviews": total_reviews,
        "total_revenue": total_revenue,

        "today_orders": today_orders,
        "pending_orders": pending_orders,
        "delivered_orders": delivered_orders,
        "cancelled_orders": cancelled_orders,

        "top_products": top_products,
        "latest_orders": latest_orders,
        "low_stock_products": low_stock_products,

        "months": json.dumps(months),
        "sales": json.dumps(sales),

         "status_labels": json.dumps(["Pending", "Delivered", "Cancelled"]),
         "status_data": json.dumps([
            pending_orders,
            delivered_orders,
            cancelled_orders
    ]),
    }

    return render(request, "adminsite/reports.html", context)

@login_required
def export_report_pdf(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="SR_Clothing_Report.pdf"'

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()
    elements = []

    # -----------------------
    # Logo
    # -----------------------
    logo_path = os.path.join(
        settings.BASE_DIR,
        "static",
        "img",
        "sr_logo.png"
    )

    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1.2*inch, height=1.2*inch)
        elements.append(logo)

    # -----------------------
    # Header
    # -----------------------

    elements.append(
        Paragraph(
            "<font size='24' color='#0d6efd'><b>SR CLOTHING</b></font>",
            styles['Title']
        )
    )

    elements.append(
        Paragraph(
            "<font size='12' color='grey'>Premium Men's Fashion Store</font>",
            styles['Heading2']
        )
    )

    elements.append(Spacer(1, 12))

    elements.append(
        Paragraph(
            "<font size='18'><b>Business Analytics Report</b></font>",
            styles['Heading1']
        )
    )

    elements.append(
        Paragraph(
            f"Generated On : {timezone.localtime().strftime('%d-%m-%Y %I:%M %p')}",
            styles['Normal']
        )
    )

    elements.append(Spacer(1, 20))

    # -----------------------
    # Statistics
    # -----------------------

    total_users = User.objects.count()
    total_products = product.objects.count()
    total_orders = Order.objects.count()
    total_payments = Payment.objects.count()
    total_reviews = Review.objects.count()

    total_revenue = (
        Order.objects.filter(status="Delivered")
        .aggregate(total=Sum("total_amount"))["total"] or 0
    )

    pending = Order.objects.filter(status="Pending").count()
    confirmed = Order.objects.filter(status="Confirmed").count()
    shipped = Order.objects.filter(status="Shipped").count()
    delivered = Order.objects.filter(status="Delivered").count()
    cancelled = Order.objects.filter(status="Cancelled").count()

    data = [

        ["Report Summary", "Value"],

        ["Total Users", total_users],

        ["Total Products", total_products],

        ["Total Orders", total_orders],

        ["Total Payments", total_payments],

        ["Total Reviews", total_reviews],

        ["Total Revenue", f"Rs. {total_revenue}"],

        ["Pending Orders", pending],

        ["Confirmed Orders", confirmed],

        ["Shipped Orders", shipped],

        ["Delivered Orders", delivered],

        ["Cancelled Orders", cancelled],

    ]

    table = Table(data, colWidths=[250,180])

    table.setStyle(TableStyle([

        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#0d6efd")),

        ('TEXTCOLOR',(0,0),(-1,0),colors.white),

        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),

        ('FONTSIZE',(0,0),(-1,0),13),

        ('BOTTOMPADDING',(0,0),(-1,0),10),

        ('GRID',(0,0),(-1,-1),0.5,colors.grey),

        ('BACKGROUND',(0,1),(-1,-1),colors.whitesmoke),

        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),

        ('ALIGN',(1,1),(-1,-1),'CENTER'),

        ('ROWBACKGROUNDS',(0,1),(-1,-1),
            [colors.white, colors.HexColor("#f4f8fc")]
        ),

    ]))

    elements.append(table)

    elements.append(Spacer(1,25))

    elements.append(
        Paragraph(
            "<b>Thank You</b><br/>Generated by <b>SR CLOTHING Management System</b>",
            styles['Heading3']
        )
    )

    doc.build(elements)

    return response

@login_required
def export_report_excel(request):

    if not request.user.is_staff:
        return HttpResponse("Access Denied!")

    wb = Workbook()
    ws = wb.active
    ws.title = "SR CLOTHING REPORT"

    # Heading
    ws['A1'] = "SR CLOTHING"
    ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(fill_type="solid", fgColor="0D6EFD")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    ws['A2'] = "Business Analytics Report"
    ws['A2'].font = Font(size=13, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:B2")

    ws.append([])

    # Statistics
    total_users = User.objects.count()
    total_products = product.objects.count()
    total_orders = Order.objects.count()
    total_payments = Payment.objects.count()
    total_reviews = Review.objects.count()

    total_revenue = (
        Order.objects.filter(status="Delivered")
        .aggregate(total=Sum("total_amount"))["total"] or 0
    )

    pending = Order.objects.filter(status="Pending").count()
    confirmed = Order.objects.filter(status="Confirmed").count()
    shipped = Order.objects.filter(status="Shipped").count()
    delivered = Order.objects.filter(status="Delivered").count()
    cancelled = Order.objects.filter(status="Cancelled").count()

    ws.append(["Report", "Value"])

    header_row = ws.max_row

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0D6EFD")
        cell.alignment = Alignment(horizontal="center")

    ws.append(["Total Users", total_users])
    ws.append(["Total Products", total_products])
    ws.append(["Total Orders", total_orders])
    ws.append(["Total Payments", total_payments])
    ws.append(["Total Reviews", total_reviews])
    ws.append(["Total Revenue", total_revenue])

    ws.append([])

    ws.append(["Pending Orders", pending])
    ws.append(["Confirmed Orders", confirmed])
    ws.append(["Shipped Orders", shipped])
    ws.append(["Delivered Orders", delivered])
    ws.append(["Cancelled Orders", cancelled])

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="SR_Clothing_Report.xlsx"'

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    wb.save(response)

    return response